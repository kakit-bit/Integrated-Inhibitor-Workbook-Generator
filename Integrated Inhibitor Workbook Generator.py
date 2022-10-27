# Script Takes User Input for Number of Samples to Run
# For Each Sample Accepts Demographics and Flavour of Inhibitor

# Script Generates Excel Documents from Respective Template
# Script Inputs Patient Demographics into Generated Document
# Script Saves Generated Document as per SOP Requirements

import openpyxl
type_values = {2:'II', 5:'V', 7:'VII', 8:'VIII', 9:'IX', 10:'X', 11:'XI', 12:'XII'}

# Class saves Patient Demographics into Profile Object
class Patient():
    def __init__(self, first_name, last_name, sample_id, inhibitor):
        self.first_name = first_name
        self.last_name = last_name
        self.sample_id = sample_id
        self.inhibitor = inhibitor

# Class saves Factor Aspects into Profile Specific Object
class Factor():
    def __init__(self, type, level, flavour, form):
        self.type = type
        self.level = level
        self.flavour = flavour # Classic, Nijmegen, Porcine
        self.form = form # Clotting, Chromogenic

# User Input Error Check Functions
def is_number(value):
    try:
        int(value)
        return True
    except ValueError:
        return False

def is_float(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

def in_range(value,limiter):
    if value not in range(1,limiter):
        return False
    else:
        return True

def is_limit(value):
    if value[0] not in (">","<"):
        return True
    elif (value[0] in (">","<")) and (len(value[1:]) > 0) and (is_float(value[1:]) == True):
        return True
    else:
        return False

# Function to Check How Many Patients to Generate
def get_workload():
    while True:
        workload_string = input("Number of Inhibitors to Run: ")
        # Check if Input is a Valid Number
        if not is_number(workload_string):
            print("Not a valid number")
            continue
        workload = int(workload_string)
        return workload

# Function Allows User to Input Patient Demographics
def get_patient():
    # Retrieve Details for Patient Class
    last_name = str(input(("Patient Last Name: "))).lower().capitalize()
    first_name = str(input(("Patient First Name: "))).lower().capitalize()
    sample_id = str(input(("Sample ID: "))).lower().capitalize()

    # Retrieve Details for Factor Class
    type = get_type()
    level = get_level(type)
    flavour = get_flavour(type)
    # Check if Nijmegen
    if flavour == "Nijmegen":
        form = get_form()
    else:
        form = None

    # Clear Line for Multiple Patients
    print("")

    # Create Factor Class
    inhibitor = Factor(type, level, flavour, form)

    # Create Patient Class
    patient = Patient(first_name, last_name, sample_id, inhibitor)

    return patient

# Function to Obtain Factor Type:
def get_type():
    while True:
        type_string = input("Enter Factor Type: ")
        # Check if Valid Number
        if not is_number(type_string):
            print("Not a Number")
            continue
        # Check if in type_value
        if int(type_string) not in type_values:
            print("No Factor Type Exists")
            continue
        else:
            # Returns Ordinal Form of Factor Number
            return type_values[int(type_string)]

# Function to Obtain Factor Result:
def get_level(type):
    while True:
        level_string = input("Factor "+type+" Result: ")
        # Check if Limit Breaker
        if not is_limit(level_string):
            # Check if Input is a Valid Number
            if not is_float(level_string):
                print("Not a valid number")
                continue
            continue
        return level_string

# Function to Obtain Inhibitor Flavour
def get_flavour(type):
    # Check if Factor Type is not FVIII
    if type != 'VIII':
        flavour = "Classic"
        return flavour
    # If Factor Type is FVIII, Provide Flavour Options
    else:
        while True:
            flavour_string = input("(1) Classic (2) Nijmegen (3) Porcine: ")
            # Check if input is Valid Number
            if not is_number(flavour_string):
                print("Not a valid number")
                continue
            flavour_input = int(flavour_string)
            if not in_range(flavour_input,4):
                print("Not a valid choice")
                continue

            if flavour_input == 1:
                flavour = "Classic"
            elif flavour_input == 2:
                flavour = "Nijmegen"
            elif flavour_input == 3:
                flavour = "Porcine"
            return flavour

# Function to Obtain Nijmegen Form:
def get_form():
    while True:
        form_string = input("(1) Clotting or (2) Chromogenic: ")
        # Check if Input is Valid Number
        if not is_number(form_string):
            print("Not a valid number")
            continue
        form_input = int(form_string)
        if not in_range(form_input,3):
            print("Not a valid choice")
            continue

        if form_input == 1:
            form = "Clotting"
        elif form_input == 2:
            form = "Chromogenic"
        return form

# Functions Opens Excel Template and Writes
def classic_excel():
    srcfile = openpyxl.load_workbook('Quantitative Classic Human Inhibitor.xlsx', read_only=False,keep_vba=False)
    sheetname = srcfile['Classic Human Inhibitor']
    sheetname['E5'] = str(patient.last_name + ', ' + patient.first_name)
    sheetname['E6'] = str(patient.sample_id)
    sheetname['C5'] = str(today_date)
    sheetname['C7'] = tech_initials
    sheetname['C9'] = patient.inhibitor.type
    sheetname['C11'] = patient.inhibitor.level

    if (patient.inhibitor.level == '>2.50'):
        sheetname['D19'] = 200
        sheetname['D21'] = 0
    elif (patient.inhibitor.level == '<0.01'):
        sheetname['D19'] = 0
        sheetname['D21'] = 200
    elif (float(patient.inhibitor.level) >= 1.00):
        sheetname['D19'] = 200
        sheetname['D21'] = 0
    elif (float(patient.inhibitor.level) <= 0.10):
        sheetname['D19'] = 0
        sheetname['D21'] = 200
    else:
        deficient_needed = float(patient.inhibitor.level) * 200
        normal_plasma = 200 - deficient_needed

        sheetname['D19'] = normal_plasma
        sheetname['D21'] = deficient_needed

    filename = patient.inhibitor.flavour + '.' + patient.sample_id + '.' + today_date + '.xlsx'
    filename = str(filename)

    srcfile.save(filename)

def nijmegen_excel():
    srcfile = openpyxl.load_workbook('FVIII Nijmegen-Bethesda Assay - Inhibitor Worksheet v2.xlsx',read_only=False, keep_vba= False)
    sheetname = srcfile['Sheet1']
    sheetname['F5'] = str(patient.last_name + ', ' + patient.first_name)
    sheetname['F6'] = str(patient.sample_id)
    sheetname['C6'] = str(today_date)
    sheetname['D8'] = str(patient.inhibitor.level)
    sheetname['D17'] = tech_initials

    if patient.inhibitor.form == "Clotting":
        sheetname['G18'] = "Y"
    else:
        sheetname['G17'] = "Y"

    filename = patient.inhibitor.flavour + patient.inhibitor.form+'.' + patient.sample_id + '.' + today_date + '.xlsx'
    filename = str(filename)

    srcfile.save(filename)

def porcine_excel():
    srcfile = openpyxl.load_workbook('Quantitative Porcine Inhibitor.xlsx', read_only=False, keep_vba=False)
    sheetname = srcfile['Porcine Inhibitor']
    sheetname['E4'] = str(patient.last_name + ', ' + patient.first_name)
    sheetname['E5'] = str(patient.sample_id)
    sheetname['C4'] = str(today_date)
    sheetname['C7'] = tech_initials
    sheetname['C9'] = patient.inhibitor.level

    filename = patient.inhibitor.flavour + '.' + patient.sample_id + '.' + today_date + '.xlsx'
    filename = str(filename)

    srcfile.save(filename)

if __name__ == '__main__':

    # Initialize Empty Patient List
    all_patients = []

    #Get Today's Date
    from time import gmtime, strftime
    today_date = strftime("%d.%m.%Y", gmtime())

    # Get User Initials
    tech_initials = str(input("Enter Your Initials: ").upper())

    # Get Workload and Patient Demographics
    num_samples = get_workload()
    for x in range(num_samples):
        patient = get_patient()
        all_patients.append(patient)

    # Code to Write Information to Respective Template
    for patient in all_patients:
        if patient.inhibitor.flavour == "Classic":
            classic_excel()
        elif patient.inhibitor.flavour == "Nijmegen":
            nijmegen_excel()
        elif patient.inhibitor.flavour == "Porcine":
            porcine_excel()